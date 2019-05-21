from openpyxl import Workbook
import pandas as pd
import numpy as np
from openpyxl.styles import Border, Side, Alignment, Font

wb = Workbook()
ws = wb.active
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
alignment = Alignment(horizontal='center')
font = Font(bold=True)

df = pd.DataFrame(np.random.rand(5, 5), columns=list('abcde'))
rows, cols = df.shape
for i in range(len(df.columns)):
    t = df.columns
    ce = ws.cell(row=3, column=i + 1)
    ce.border = border
    ce.alignment = alignment
    ce.font = font
    ce.value = t[i]


for r in range(rows):
    for c in range(cols):
        ws.cell(row=3 + r + 1, column=c + 1,
                value=df.iloc[r, c]).border = border
wb.save("test.xltx")
